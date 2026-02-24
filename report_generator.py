from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Font, Alignment
from backend import compute_layer_vsi

TEMPLATE_PATH = "Sample.xlsx"
OUTPUT_PATH = "Site_Class_Report.xlsx"


def build_formula_text(breakdown_layer):

    soil = breakdown_layer["soil_type"]
    n1 = breakdown_layer["n1"]
    fines = breakdown_layer["fines"]

    if soil == "Others" or n1 < 10:
        return "NA"

    if soil == "Dry Sands":
        exponent = "⁰·⁵" if fines == "Yes" else "⁰·³"
    elif soil == "Saturated Sands":
        exponent = "⁰·⁴" if fines == "Yes" else "⁰·³"
    elif soil == "Clays":
        exponent = "⁰·³"
    else:
        return "NA"

    return f"80[(N₁)₆₀ᵢ]{exponent}"


def generate_site_class_report(site_data, result):

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Site Class Report"]

    layers = result["breakdown"]
    n_layer = len(layers)

    start_idx = 11
    max_rows = 20   # fixed capacity

    # ----------------------------
    # Top Summary
    # ----------------------------

    ws["B3"] = "Weighted Average Shear Wave Velocity Vₛ ="
    ws["N3"] = round(result["weighted_vs"], 3)
    ws["G4"] = result["site_class"]
    ws["H7"] = site_data["depth_of_influence"]

    # ----------------------------
    # Clear Previous Layer Area
    # ----------------------------

    for i in range(max_rows):
        row = start_idx + i
        for col in [2, 5, 9, 14, 17, 20, 25, 28]:
            ws.cell(row=row, column=col).value = None

    ws["E9"] = "Thickmess (tᵢ)"
    ws["Q9"] = "(N₁)₆₀ᵢ"
    ws["T9"] = "Shear Wave Velocity Vₛᵢ"
    ws["AB9"] = "tᵢ/Vₛᵢ"

    # ----------------------------
    # Fill Layer Data
    # ----------------------------

    for i, layer in enumerate(layers):

        row = start_idx + i
        ti = layer["effective_thickness"]
        vsi = layer["computed_vsi"]
        ti_over_vsi = layer["ti_over_vsi"]

        ws.cell(row=row, column=2).value = f"Layer {i+1}"
        ws.cell(row=row, column=5).value = round(ti, 3)
        ws.cell(row=row, column=9).value = layer["soil_type"]
        if layer["soil_type"] in ["Clays", "Others"]:
            ws.cell(row=row, column=14).value = "NA"
        else:
            ws.cell(row=row, column=14).value = layer["fines"]
        ws.cell(row=row, column=17).value = layer["n1"]
        formula_text = build_formula_text(layer)
        ws.cell(row=row, column=20).value = formula_text
        ws.cell(row=row, column=25).value = round(vsi, 3)
        ws.cell(row=row, column=28).value = round(ti_over_vsi, 6)

    # ----------------------------
    # Σ Row
    # ----------------------------

    sum_row = start_idx + n_layer

    total_ti = sum(layer["effective_thickness"] for layer in layers)
    total_ti_over_vsi = sum(layer["ti_over_vsi"] for layer in layers)

    ws.cell(row=sum_row, column=2).value = "Σtᵢ"
    ws.cell(row=sum_row, column=5).value = round(total_ti, 3)
    ws.cell(row=sum_row, column=25).value = "Σ(tᵢ/Vₛᵢ)"
    ws.cell(row=sum_row, column=28).value = round(total_ti_over_vsi, 6)

    
     # ----------------------------
    # Clear Below Σ Area
    # ----------------------------

    for i in range(sum_row + 1, 40): 
        for col in [2, 5, 9, 14, 17, 20, 25, 28]:
            ws.cell(row=i, column=col).value = None
        for j in range(16):
            ws.cell(row=sum_row, column=9+j).border = Border()  
        for col in range(2,33):
            ws.cell(row=i, column=col).border = Border()

    # ----------------------------
    # Clear Previous Highlight
    # ----------------------------

    class_rows = {
        "A": 11,
        "B": 12,
        "C": 13,
        "D": 14,
        "E": 15
    }

    for r in class_rows.values():
        ws.cell(row=r, column=33).fill = PatternFill(fill_type=None)
        ws.cell(row=r, column=36).fill = PatternFill(fill_type=None)

    # ----------------------------
    # Apply Highlight
    # ----------------------------

    site_class = result["site_class"]
    highlight_row = class_rows.get(site_class)

    if highlight_row:
        fill = PatternFill(
            start_color="DAEEF3",
            end_color="DAEEF3",
            fill_type="solid"
        )
        ws.cell(row=highlight_row, column=33).fill = fill
        ws.cell(row=highlight_row, column=36).fill = fill
    
    ws["AG17"] = "Vₛ = Σtᵢ / Σ(tᵢ / Vₛᵢ) = "
    ws["AL17"] = "=N3"
    ws["AM18"] = "=G4"

    ws["AJ11"] =  "Vₛ ≥ 1500"
    ws["AJ12"] =  "760 ≤ Vₛ < 1500"
    ws["AJ13"] =  "360 ≤ Vₛ < 760"
    ws["AJ14"] =  "180 ≤ Vₛ < 360"
    ws["AJ15"] =  "Vₛ ≤ 180"


   # ----------------------------
    # Add Note
    # ----------------------------

    note_row = sum_row + 3
    print(note_row)

    note_text = "Note - For last layer, the thickness as required based on the Depth of Influence is considered."

    
    # Merge across full table width (adjust if your table goes wider)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=2, end_column=4)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=5, end_column=8)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=9, end_column=13)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=14, end_column=16)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=17, end_column=19)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=20, end_column=24)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=25, end_column=27)
    ws.unmerge_cells(start_row=note_row, end_row=note_row, start_column=28, end_column=31)

    ws.merge_cells(start_row=note_row, end_row=note_row, start_column=2, end_column=31)

    ws.cell(row=note_row, column=2).value = note_text

    # Enable wrapping + top alignment
    ws.cell(row=note_row, column=2).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True
    )

    # Optional styling
    ws.cell(row=note_row, column=2).font = Font(italic=True)

    # Increase row height so text is fully visible
    # ws.row_dimensions[note_row].height = 30


    # ----------------------------
    # Save
    # ----------------------------

    wb.save(OUTPUT_PATH)
    wb.close()
    return OUTPUT_PATH
